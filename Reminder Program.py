import openpyxl,time
from datetime import datetime


while True:
	wb = openpyxl.load_workbook('tasks.xlsx')
	sheet = wb.active
	for i in range(2,sheet.max_row+1):
		d = sheet.cell(row=i,column=5).value
		e = sheet.cell(row=i,column=4).value
		if datetime.now().strftime('%H:%M:%S')==str(d) and datetime.now().strftime('%Y-%m-%d 00:00:00')==str(e):
			from tkinter import *
			root = Tk()
			root.geometry("350x150")
			root.title('Reminder')
			a = Label(root,text=sheet.cell(row=i,column=2).value,font=('courier',19,'bold','underline'),fg='blue')
			a.pack(anchor='w')
			b = Label(root,text='\n\n'+sheet.cell(row=i,column=3).value,font=('courier',12))
			b.pack(anchor='w')
			root.mainloop()
		